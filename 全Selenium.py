# -*- coding: utf-8 -*- 
# @email : upapqqxyz@gmail.com
# @Author : Ane
from selenium import webdriver
from lxml import etree
import time
import os
from wordcloud import WordCloud
import openpyxl
wb=webdriver.Chrome("E:/WebDriver/chromedriver.exe")
def getpage(num=1):
    wb.get(f"https://www.zhipin.com/c101010100-p100109/?page={num}")
    time.sleep(1)
    return wb.page_source
def parserhtml(html):
    tree=etree.HTML(html)
    ul=tree.xpath('/html/body/div[1]/div[3]/div/div[2]/ul/li')
    if ul==[]:
        ul=tree.xpath('/html/body/div[1]/div[3]/div/div[3]/ul/li')
        if ul ==[]:
         return False
    joblist=[]
    for li in ul:
        job = {}
        job['name']=li.xpath("./div/div[1]/div[1]/div/div[1]/span[1]/a/text()")[0]  #工作名称
        job['money']=li.xpath("./div/div[1]/div[1]/div/div[2]/span/text()")[0]  #工资
        job['exp']=li.xpath("./div/div[1]/div[1]/div/div[2]/p//text()")[0]      #经验
        job['edu']=li.xpath("./div/div[1]/div[1]/div/div[2]/p//text()")[1]      #学历
        text="".join(li.xpath("./div/div[2]/div[1]//text()"))               #技术栈
        job['ability'] =text.replace("\n","").replace("                                                    ",",").replace("                                        ","")[1:]
        job['address']=li.xpath("./div/div[1]/div[1]/div/div[1]/span[2]/span/text()")[0]  #工作地点
        job['company']=li.xpath("./div/div[1]/div[2]/div/h3/a/text()")[0]  #企业名称
        text=",".join(li.xpath("./div/div[1]/div[2]/div/p//text()"))
        job['cminfo']=text          #企业信息
        text=li.xpath("./div/div[2]/div[2]/text()")
        if text == []:    #员工福利(部分企业存在员工福利一栏为空的情况,加以判断防止报错)
            job['gift']=""
        else:
            job['gift']=text[0]
        joblist.append(job)
    return joblist
class Excel:
    x = 1
    y = 1
    if os.path.exists("Data.xlsx") == False:
        wb = openpyxl.Workbook()
        ws = wb.active
        biaoti = ["工作名称", "工资", "经验程度", "学历", "技术栈", "工作地点", "企业名称", "企业信息", "员工福利"]
        for neirong in biaoti:
            ws.cell(y, x).value = neirong
            x += 1
        y += 1
        x = 1
    else:
        wb = openpyxl.load_workbook("Data.xlsx")
        ws = wb.active
        while ws.cell(y,x).value != None:
            y+=1
    def check(self):
        return self.y,self.x
    def write(self,text,hang=y,lie=x):
        self.ws.cell(hang,lie).value=text
    def save(self):
        self.wb.save("Data.xlsx")
        self.wb.close()

def creatwordcloud(text):
    text = text.replace(",", " ")
    wordclud = WordCloud(scale=5, background_color="White", font_path="SIMYOU.TTF", relative_scaling=0.1).generate(text)
    wordclud.to_file("test.png")
if __name__ == "__main__":
    num=1
    words=""
    excel=Excel()
    y,x=excel.check()
    while num<=8:
        try:
            wb.get(f"https://www.zhipin.com/c101010100-p100109/?page={num}&ka=page-{num}")
            time.sleep(1)
            html = wb.page_source
            joblist = parserhtml(html)
            for job in joblist:
                for text in job:
                    excel.write(job[text],y,x)
                    if text=="ability":
                        words=words+job[text]+","
                    x+=1
                y += 1
                x = 1
            num+=1
        except Exception as error:
            print(error)
            with open('error.html',"w",encoding="utf-8") as file:
                file.write(wb.page_source)
    print(words)
    creatwordcloud(words)
    excel.save()
    wb.quit()